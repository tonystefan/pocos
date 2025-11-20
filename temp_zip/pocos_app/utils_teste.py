import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_intervalos_tempo(tempo_total_horas, tempo_estabilizacao_min):
    """
    Gera a lista de intervalos de tempo (em minutos) para as leituras,
    seguindo as regras específicas do usuário.
    """
    intervalos = []
    
    # 1. Primeira hora (60 minutos)
    # 3 primeiras a cada 1 minuto
    intervalos.extend([1, 1, 1])
    # próximas 3 a cada 5 minutos
    intervalos.extend([5, 5, 5])
    # 4 demais a cada 10 minutos
    intervalos.extend([10, 10, 10, 10])
    
    tempo_decorrido_min = sum(intervalos)
    
    # 2. A partir da segunda hora até o tempo de estabilização
    tempo_estabilizacao_total_min = tempo_estabilizacao_min
    
    # Intervalo de 15 minutos
    while tempo_decorrido_min < tempo_estabilizacao_total_min:
        intervalo = 15
        if tempo_decorrido_min + intervalo > tempo_estabilizacao_total_min:
            intervalo = tempo_estabilizacao_total_min - tempo_decorrido_min
        
        if intervalo > 0:
            intervalos.append(intervalo)
            tempo_decorrido_min += intervalo
        else:
            break
            
    # 3. Após a estabilização até o fim do tempo total
    tempo_total_min = tempo_total_horas * 60
    
    # Intervalo de 60 minutos (1 hora)
    while tempo_decorrido_min < tempo_total_min:
        intervalo = 60
        if tempo_decorrido_min + intervalo > tempo_total_min:
            intervalo = tempo_total_min - tempo_decorrido_min
            
        if intervalo > 0:
            intervalos.append(intervalo)
            tempo_decorrido_min += intervalo
        else:
            break
            
    return intervalos

def simular_progresso_randomico(num_leituras, valor_inicial, valor_final):
    """
    Gera uma série de valores progressivos e randômicos entre valor_inicial e valor_final.
    """
    if num_leituras <= 1:
        return [valor_final]
        
    # Simular progressão randômica entre valor_inicial e valor_final
    # Vazão: Vazão inicial > Vazão final (Progressão decrescente)
    # Nível: Nível inicial < Nível final (Progressão crescente)
    
    # Se o valor final for menor que o inicial, a progressão é decrescente (ex: Vazão)
    if valor_inicial > valor_final:
        diferenca = valor_inicial - valor_final
        sinal = -1
    else:
        diferenca = valor_final - valor_inicial
        sinal = 1
        
    # Gera num_leituras - 1 passos aleatórios que somam 1 (para garantir a progressão total)
    passos_relativos = np.random.rand(num_leituras - 1)
    passos_relativos /= passos_relativos.sum()
    
    # Converte para passos absolutos
    passos_absolutos = passos_relativos * diferenca * sinal
    
    # Calcula os valores progressivos
    valores = [valor_inicial]
    valor_atual = valor_inicial
    for passo in passos_absolutos:
        # Adiciona um pequeno ruído randômico para simular a leitura
        ruido = np.random.uniform(-0.01, 0.01) # Ruído de +/- 0.01
        valor_atual += passo + ruido
        valores.append(valor_atual)
        
    # Garante que o último valor seja exatamente o valor_final
    valores[-1] = valor_final
    
    return valores

def gerar_teste_bombeamento(params_bombeamento):
    """
    Gera o DataFrame para o Teste de Bombeamento.
    """
    data_hora_inicial = datetime.combine(params_bombeamento['data_inicio'], params_bombeamento['hora_inicial'])
    nivel_inicial = params_bombeamento['nivel_inicial']
    nivel_final = params_bombeamento['nivel_final']
    vazao_inicial = params_bombeamento['vazao_inicial']
    vazao_final = params_bombeamento['vazao_final']
    tempo_estabilizacao_min = params_bombeamento['tempo_estabilizacao_min']
    tempo_total_horas = params_bombeamento['tempo_total_horas']
    
    # 1. Gerar os intervalos de tempo
    intervalos_min = gerar_intervalos_tempo(tempo_total_horas, tempo_estabilizacao_min)
    
    # 2. Inicializar listas de dados
    leituras = []
    
    # 3. Primeira linha (Hora 0)
    leituras.append({
        'Tempo (min)': 0,
        'Data/Hora': data_hora_inicial,
        'Nível (m)': nivel_inicial,
        'Vazão (m³/h)': 0.0, # Vazão é 0 na primeira leitura
    })
    
    # 4. Simular progressão até o tempo de estabilização
    
    # Encontrar o índice da leitura que atinge ou ultrapassa o tempo de estabilização
    tempo_acumulado = 0
    indice_estabilizacao = -1
    for i, intervalo in enumerate(intervalos_min):
        tempo_acumulado += intervalo
        if tempo_acumulado >= tempo_estabilizacao_min:
            indice_estabilizacao = i
            break
            
    # Número de leituras até a estabilização (incluindo a leitura de estabilização)
    num_leituras_progressivas = indice_estabilizacao + 1
    
# Simular a progressão do Nível (de nivel_inicial a nivel_final)
# A primeira leitura de nível é nivel_inicial, então a progressão deve começar na segunda leitura.
    niveis_progressivos_completo = simular_progresso_randomico(num_leituras_progressivas + 1, nivel_inicial, nivel_final)
    niveis_progressivos = niveis_progressivos_completo[1:] # Exclui o nivel_inicial (que já está na linha 0)
    
    # Simular a progressão da Vazão (de vazao_inicial a vazao_final)
    # A primeira vazão é 0, a segunda é vazao_inicial, então precisamos de num_leituras_progressivas - 1 passos
    vazoes_progressivas = simular_progresso_randomico(num_leituras_progressivas, vazao_inicial, vazao_final)
    
    # 5. Preencher as leituras
    tempo_acumulado = 0
    data_hora_atual = data_hora_inicial
    
    for i, intervalo in enumerate(intervalos_min):
        tempo_acumulado += intervalo
        data_hora_atual += timedelta(minutes=intervalo)
        
        leitura = {
            'Tempo (min)': tempo_acumulado,
            'Data/Hora': data_hora_atual,
        }
        
        if i < num_leituras_progressivas:
            # Leituras progressivas
            leitura['Nível (m)'] = niveis_progressivos[i]
            leitura['Vazão (m³/h)'] = vazoes_progressivas[i]
        else:
            # Leituras após a estabilização (repetem o valor final)
            leitura['Nível (m)'] = nivel_final
            leitura['Vazão (m³/h)'] = vazao_final
            
        leituras.append(leitura)
        
    # Ajustar a segunda linha para ter vazao_inicial
    if len(leituras) > 1:
        leituras[1]['Vazão (m³/h)'] = vazao_inicial
        
    df = pd.DataFrame(leituras)
    
    # Formatação: 1 casa decimal
    df['Nível (m)'] = df['Nível (m)'].round(1)
    if 'Vazão (m³/h)' in df.columns:
        df['Vazão (m³/h)'] = df['Vazão (m³/h)'].round(1)
        
    # Remover a coluna de data (manter apenas a hora no Data/Hora)
    # A coluna Data/Hora será convertida para string na view, mas aqui vamos garantir que não haja data
    df['Data/Hora'] = df['Data/Hora'].dt.strftime('%H:%M:%S')
    
    return df

def gerar_teste_recuperacao(params_recuperacao, nivel_inicial_rec, nivel_final_rec):
    """
    Gera o DataFrame para o Teste de Recuperação.
    """
    data_hora_inicial = datetime.combine(params_recuperacao['data_inicio'], params_recuperacao['hora_inicial'])
    data_hora_final = datetime.combine(params_recuperacao['data_inicio'], params_recuperacao['hora_final'])
    tempo_estabilizacao_min = params_recuperacao['tempo_estabilizacao_min']
    num_leituras = params_recuperacao['num_leituras']
    
    # 1. Inicializar listas de dados
    leituras = []
    
    # 2. Primeira linha (Hora 0)
    leituras.append({
        'Tempo (min)': 0,
        'Data/Hora': data_hora_inicial,
        'Nível (m)': nivel_inicial_rec,
    })
    
    # 3. Simular progressão até o tempo de estabilização
    
    # Calcular o intervalo de tempo para as leituras progressivas
    intervalo_progressivo = tempo_estabilizacao_min / num_leituras
    
# Simular a progressão do Nível (de nivel_inicial_rec a nivel_final_rec)
# A primeira leitura de nível é nivel_inicial_rec, então a progressão deve começar na segunda leitura.
    niveis_progressivos_completo = simular_progresso_randomico(num_leituras + 1, nivel_inicial_rec, nivel_final_rec)
    niveis_progressivos = niveis_progressivos_completo[1:] # Exclui o nivel_inicial_rec (que já está na linha 0)
    
    tempo_acumulado = 0
    data_hora_atual = data_hora_inicial
    
    for i in range(num_leituras):
        tempo_acumulado += intervalo_progressivo
        data_hora_atual += timedelta(minutes=intervalo_progressivo)
        
        leitura = {
            'Tempo (min)': tempo_acumulado,
            'Data/Hora': data_hora_atual,
            'Nível (m)': niveis_progressivos[i],
        }
        leituras.append(leitura)
        
    # 4. Leituras após a estabilização até o fim do teste
    
    # O tempo total do teste de recuperação é a diferença entre hora_final e hora_inicial
    tempo_total_recuperacao_min = (data_hora_final - data_hora_inicial).total_seconds() / 60
    
    # Intervalo de 60 minutos (1 hora) após a estabilização
    tempo_decorrido_min = tempo_acumulado
    
    while tempo_decorrido_min < tempo_total_recuperacao_min:
        intervalo = 60
        if tempo_decorrido_min + intervalo > tempo_total_recuperacao_min:
            intervalo = tempo_total_recuperacao_min - tempo_decorrido_min
            
        if intervalo > 0:
            tempo_decorrido_min += intervalo
            data_hora_atual += timedelta(minutes=intervalo)
            
            leitura = {
                'Tempo (min)': tempo_decorrido_min,
                'Data/Hora': data_hora_atual,
                'Nível (m)': nivel_final_rec, # Repete o valor final
            }
            leituras.append(leitura)
        else:
            break
            
    df = pd.DataFrame(leituras)
    
    # Formatação: 1 casa decimal
    df['Nível (m)'] = df['Nível (m)'].round(1)
    if 'Vazão (m³/h)' in df.columns:
        df['Vazão (m³/h)'] = df['Vazão (m³/h)'].round(1)
        
    # Remover a coluna de data (manter apenas a hora no Data/Hora)
    # A coluna Data/Hora será convertida para string na view, mas aqui vamos garantir que não haja data
    df['Data/Hora'] = df['Data/Hora'].dt.strftime('%H:%M:%S')
    
    return df

def gerar_teste_xlsx_file(df_bombeamento, df_recuperacao, params_bombeamento, params_recuperacao, nivel_inicial_rec, nivel_final_rec):
    """
    Gera o arquivo XLSX com os DataFrames de Teste de Bombeamento e Recuperação.
    """
    output = io.BytesIO()
    wb = Workbook()
    
    # --- Estilos ---
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # --- Teste de Bombeamento ---
    ws_bombeamento = wb.active
    ws_bombeamento.title = "Teste de Bombeamento"
    
    # Adicionar cabeçalho de parâmetros
    ws_bombeamento.append(["Parâmetros do Teste de Bombeamento"])
    ws_bombeamento.merge_cells('A1:B1')
    ws_bombeamento['A1'].font = font_bold
    
    parametros_bombeamento = [
        ("Data de Início", params_bombeamento['data_inicio'].strftime('%d/%m/%Y')),
        ("Hora de Início", params_bombeamento['hora_inicial'].strftime('%H:%M:%S')),
        ("Nível Inicial (m)", params_bombeamento['nivel_inicial']),
        ("Nível Final (m)", params_bombeamento['nivel_final']),
        ("Vazão Inicial (m³/h)", params_bombeamento['vazao_inicial']),
        ("Vazão Final (m³/h)", params_bombeamento['vazao_final']),
        ("Tempo de Estabilização (min)", params_bombeamento['tempo_estabilizacao_min']),
        ("Tempo Total do Teste (h)", params_bombeamento['tempo_total_horas']),
    ]
    
    for row_data in parametros_bombeamento:
        ws_bombeamento.append(row_data)
        
    # Adicionar uma linha em branco
    ws_bombeamento.append([])
    
    # Adicionar o DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df_bombeamento, header=True, index=False)):
        ws_bombeamento.append(row)
        
    # Aplicar estilos ao cabeçalho da tabela
    header_row = ws_bombeamento[ws_bombeamento.max_row - len(df_bombeamento) - 1]
    for cell in header_row:
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        
    # Ajustar largura das colunas
    for i, col in enumerate(ws_bombeamento.columns):
        max_length = 0
        column = get_column_letter(i + 1) # Usa o índice para obter a letra da coluna
        for cell in col:
            cell_value_str = str(cell.value) if cell.value is not None else ""
            if len(cell_value_str) > max_length:
                max_length = len(cell_value_str)
        adjusted_width = (max_length + 2)
        ws_bombeamento.column_dimensions[column].width = adjusted_width
        
    # --- Teste de Recuperação ---
    ws_recuperacao = wb.create_sheet(title="Teste de Recuperação")
    
    # Adicionar cabeçalho de parâmetros
    ws_recuperacao.append(["Parâmetros do Teste de Recuperação"])
    ws_recuperacao.merge_cells('A1:B1')
    ws_recuperacao['A1'].font = font_bold
    
    parametros_recuperacao = [
        ("Data de Início", params_recuperacao['data_inicio'].strftime('%d/%m/%Y')),
        ("Hora de Início", params_recuperacao['hora_inicial'].strftime('%H:%M:%S')),
        ("Hora de Fim", params_recuperacao['hora_final'].strftime('%H:%M:%S')),
        ("Tempo de Estabilização (min)", params_recuperacao['tempo_estabilizacao_min']),
        ("Número de Leituras até o Nível Final", params_recuperacao['num_leituras']),
        ("Nível Inicial (m)", nivel_inicial_rec),
        ("Nível Final (m)", nivel_final_rec),
    ]
    
    for row_data in parametros_recuperacao:
        ws_recuperacao.append(row_data)
        
    # Adicionar uma linha em branco
    ws_recuperacao.append([])
    
    # Adicionar o DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df_recuperacao, header=True, index=False)):
        ws_recuperacao.append(row)
        
    # Aplicar estilos ao cabeçalho da tabela
    header_row = ws_recuperacao[ws_recuperacao.max_row - len(df_recuperacao) - 1]
    for cell in header_row:
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        
    # Ajustar largura das colunas
    for i, col in enumerate(ws_recuperacao.columns):
        max_length = 0
        column = get_column_letter(i + 1) # Usa o índice para obter a letra da coluna
        for cell in col:
            cell_value_str = str(cell.value) if cell.value is not None else ""
            if len(cell_value_str) > max_length:
                max_length = len(cell_value_str)
        adjusted_width = (max_length + 2)
        ws_recuperacao.column_dimensions[column].width = adjusted_width
        
    wb.save(output)
    output.seek(0)
    return output.getvalue()
