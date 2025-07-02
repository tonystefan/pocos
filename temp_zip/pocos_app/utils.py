import random
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time
import calendar
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter

def gerar_datas_aleatorias(ano_inicio, ano_fim, dias_por_mes):
    """
    Gera datas aleatórias distribuídas conforme os dias por mês especificados.
    
    Args:
        ano_inicio: Ano inicial
        ano_fim: Ano final
        dias_por_mes: Dicionário com a quantidade de dias para cada mês
        
    Returns:
        Lista de datas ordenadas
    """
    datas = []
    
    # Converter códigos de mês para números
    meses_map = {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }
    
    # Para cada ano no intervalo
    for ano in range(ano_inicio, ano_fim + 1):
        # Para cada mês
        for codigo_mes, num_mes in meses_map.items():
            # Pular se não houver dias para este mês
            """dias_a_selecionar_param = dias_por_mes.get(f'dias_{codigo_mes}', 0)
            if not isinstance(dias_a_selecionar_param, (int, float)) or dias_a_selecionar_param <= 0:
                continue
            dias_a_selecionar_param = int(dias_a_selecionar_param)"""
                
            # Obter o número de dias no mês
            try:
                dias_no_mes = calendar.monthrange(ano, num_mes)[1]
            except ValueError:
                continue # Mês inválido
            
            # Número de dias a selecionar para este mês
            #dias_a_selecionar = min(dias_a_selecionar_param, dias_no_mes)
            dias_a_selecionar = dias_no_mes
            if dias_a_selecionar <= 0:
                continue
            
            # Selecionar dias aleatórios
            try:
                dias_selecionados = sorted(random.sample(range(1, dias_no_mes + 1), dias_a_selecionar))
            except ValueError:
                # Caso dias_a_selecionar seja maior que dias_no_mes (não deveria acontecer com min)
                continue 
            
            # Adicionar datas à lista
            for dia in dias_selecionados:
                try:
                    datas.append(datetime(ano, num_mes, dia))
                except ValueError:
                    continue # Dia inválido para o mês
    
    # Ordenar datas
    datas.sort()
    
    return datas

def distribuir_valores(total, num_valores, max_valor):
    """
    Distribui um valor total em um número específico de valores aleatórios,
    respeitando um valor máximo por item.
    
    Args:
        total: Valor total a ser distribuído
        num_valores: Número de valores a gerar
        max_valor: Valor máximo por item
        
    Returns:
        Lista de valores que somam o total
    """
    if num_valores <= 0:
        return []
        
    if max_valor <= 0:
         # Se max_valor for zero ou negativo, distribuir igualmente
         if num_valores > 0:
             return [total / num_valores] * num_valores
         else:
             return []

    if max_valor * num_valores < total:
        # Se não for possível distribuir, retorna valores proporcionais até o máximo
        print(f"Aviso: Impossível distribuir {total} em {num_valores} valores com máximo de {max_valor}. Distribuindo proporcionalmente até o máximo.")
        valor_medio = total / num_valores
        if valor_medio > max_valor:
             return [max_valor] * num_valores
        # Se o valor médio for menor que o máximo, mas ainda assim a soma extrapolar
        # Isso não deveria acontecer pela condição inicial, mas por segurança:
        return [min(valor_medio, max_valor)] * num_valores

    # Tentar gerar valores aleatórios que somem o total
    valores = np.random.rand(num_valores)
    valores = valores / np.sum(valores) * total
    
    # Ajustar valores que excedem o máximo
    excesso = 0
    for i in range(num_valores):
        if valores[i] > max_valor:
            excesso += valores[i] - max_valor
            valores[i] = max_valor
            
    # Redistribuir o excesso entre os valores abaixo do máximo
    indices_abaixo_max = [i for i, v in enumerate(valores) if v < max_valor]
    while excesso > 1e-6 and indices_abaixo_max:
        distribuir_por_indice = excesso / len(indices_abaixo_max)
        excesso_restante_iter = 0
        novos_indices_abaixo_max = []
        for i in indices_abaixo_max:
            adicao = min(distribuir_por_indice, max_valor - valores[i])
            valores[i] += adicao
            excesso_restante_iter += distribuir_por_indice - adicao
            if valores[i] < max_valor:
                 novos_indices_abaixo_max.append(i)
        excesso = excesso_restante_iter
        indices_abaixo_max = novos_indices_abaixo_max
        if not indices_abaixo_max and excesso > 1e-6:
             # Se não há mais onde distribuir e ainda há excesso, pode indicar problema
             # print(f"Aviso: Não foi possível redistribuir todo o excesso ({excesso:.4f}). A soma pode não ser exata.")
             break # Evita loop infinito

    # Ajuste final para garantir a soma exata devido a arredondamentos
    soma_atual = np.sum(valores)
    diferenca = total - soma_atual
    if abs(diferenca) > 1e-6:
        # Adicionar a diferença a um valor que não exceda o máximo
        idx_ajuste = -1
        for i in range(num_valores):
             if valores[i] + diferenca <= max_valor and valores[i] + diferenca >= 0:
                  idx_ajuste = i
                  break
        if idx_ajuste != -1:
             valores[idx_ajuste] += diferenca
        # else: # Se não for possível ajustar sem violar o máximo, a soma pode ficar ligeiramente diferente
            # print(f"Aviso: Não foi possível ajustar a diferença final ({diferenca:.4f}) sem violar o máximo. Soma final: {np.sum(valores):.4f}")

    return [round(v, 3) for v in valores]

def gerar_tabela_dados(parametros):
    """
    Gera a tabela de dados com base nos parâmetros fornecidos.
    
    Args:
        parametros: Dicionário com os parâmetros do formulário
        
    Returns:
        DataFrame pandas com os dados gerados
    """
    # Extrair parâmetros
    ano_inicio = parametros['ano_inicio']
    ano_fim = parametros['ano_fim']
    horimetro_inicial = parametros['horimetro_inicial']
    horimetro_final = parametros['horimetro_final']
    hidrometro_inicial = parametros['hidrometro_inicial']
    hidrometro_final = parametros['hidrometro_final']
    max_horimetro_diario = parametros['max_horimetro_diario']
    max_hidrometro_diario = parametros['max_hidrometro_diario']
    
    # Gerar datas aleatórias
    datas_dt = gerar_datas_aleatorias(ano_inicio, ano_fim, parametros)
    if not datas_dt:
        return pd.DataFrame() # Retorna DataFrame vazio se não houver datas
        
    datas_str = [d.strftime('%d/%m/%Y') for d in datas_dt]
    num_datas = len(datas_dt)
    
    # Calcular diferenças totais
    diferenca_horimetro = horimetro_final - horimetro_inicial
    diferenca_hidrometro = hidrometro_final - hidrometro_inicial
    
    # Distribuir valores diários de horímetro e hidrômetro
    valores_horimetro_diario = distribuir_valores(diferenca_horimetro, num_datas, max_horimetro_diario)
    valores_hidrometro_diario = distribuir_valores(diferenca_hidrometro, num_datas, max_hidrometro_diario)
    
    # Calcular valores acumulados
    horimetro_acumulado = np.cumsum(valores_horimetro_diario) + horimetro_inicial
    hidrometro_acumulado = np.cumsum(valores_hidrometro_diario) + hidrometro_inicial
    
    # Arredondar para 3 casas decimais
    horimetro_acumulado = [round(v, 3) for v in horimetro_acumulado]
    hidrometro_acumulado = [round(v, 3) for v in hidrometro_acumulado]
    
    # Obter vazões e horas/dia para cada data com base no mês
    
    horas_dia = [] # Corresponde a 'Tempo de Captação (h)'
    meses_map_rev = {v: k for k, v in {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }.items()}
    
    volume_acumulado_mensal = {}
    volume_diario_list = []
    tempo_diario_list = []
    volume_acum_mensal_list = []
    vazoes = []
    hora_list = []
    
    for i, data_dt in enumerate(datas_dt):
        mes_num = data_dt.month
        mes_codigo = meses_map_rev.get(mes_num, 'jan')
        
        hora = 8
        minuto = random.randint(10, 59)        
        horario_aleatorio = (f"{hora}:{minuto}")
        hora_list.append(horario_aleatorio)
        vazao_mes = parametros.get(f'vazao_{mes_codigo}', 0.0)
        horas_dia_mes = parametros.get(f'horas_{mes_codigo}', 0.0)
        
        
        horas_dia.append(horas_dia_mes)
        
        # Calcular Volume diário (m3) - Usando a diferença do hidrômetro diário
        volume_diario = valores_hidrometro_diario[i]
        tempo_diario = valores_horimetro_diario[i]
        volume_diario_list.append(round(volume_diario, 3))
        tempo_diario_list.append(round(tempo_diario, 3))

        vazoes.append(round((volume_diario / tempo_diario),2))

        # Calcular Volume acumulado Mensal (m3)
        chave_mes_ano = data_dt.strftime('%Y-%m')
        volume_acumulado_mensal[chave_mes_ano] = volume_acumulado_mensal.get(chave_mes_ano, 0) + volume_diario
        volume_acum_mensal_list.append(round(volume_acumulado_mensal[chave_mes_ano], 3))

    # Criar DataFrame com as colunas na ordem desejada pelo template
    df = pd.DataFrame({
        'Data': datas_str,
        'Hora': hora_list,
        'Horimetro': horimetro_acumulado,
        'Medidor de Vazão': hidrometro_acumulado, # Mapeando Hidrômetro para Medidor de Vazão
        'Tempo de Captação (h)': tempo_diario_list, # Mapeando Horas/Dia
        'Volume diário (m3)': volume_diario_list,
        'Volume acumulado Mensal (m3)': volume_acum_mensal_list,
        'Valor': vazoes, # Mapeando Vazão m³/h para Valor (Vazão Média - diária)
        'Unidade': ['m³/h'] * num_datas, # Unidade para Vazão Média
        'Observação': [''] * num_datas # Coluna 'Observação' vazia
    })
    
    return df

def exportar_para_xlsx(df, parametros):
    """
    Exporta o DataFrame para um arquivo XLSX em memória, seguindo o modelo.
    
    Args:
        df: DataFrame pandas com os dados
        parametros: Dicionário com os parâmetros do formulário (para cabeçalho)
        
    Returns:
        Bytes do arquivo XLSX
    """
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Planilha1'

    sheet.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )


    # --- Estilos --- (Opcional, mas melhora a aparência)
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0 
    sheet.freeze_panes = 'A7'
    sheet.sheet_view.showGridLines = False
    header_font = Font(name='Calibri', size=11, bold=True)
    title_font = Font(name='Calibri', size=11, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    fill_cinza = PatternFill(fill_type='solid', fgColor='D9D9D9')

    # --- Cabeçalho Complexo --- 
    # Linha 1: Título principal
    sheet.merge_cells('A1:J1')
    sheet['A1'] = 'PLANILHA DE MONITORAMENTO DE VAZÃO'
    sheet['A1'].font = title_font
    sheet['A1'].alignment = center_alignment

    # Linha 2: Portaria e versão
    sheet.merge_cells('A2:B2')
    sheet.merge_cells('C2:H2')
    sheet['A2'] = 'Portaria'
    sheet['I2'] = 'Versão'
    sheet['J2'] = '2020.01' # Ou buscar de algum lugar se for dinâmico
    for col in ['A', 'B', 'C', 'D','E','F','G','H','I','J']:
        cell = sheet[f'{col}2']
        cell.border = thin_border

    # Adicionar valor da portaria se disponível nos parâmetros
    # sheet['I2'] = parametros.get('portaria', '') 

    # Linha 3: Tipo de Medidor
    sheet.merge_cells('C3:J3')
    sheet['A3'] = 'Tipo de Medidor'
    for col in ['A', 'B', 'C', 'D','E','F','G','H','I','J']:
        cell = sheet[f'{col}3']
        cell.border = thin_border

    # Adicionar valor se disponível
    # sheet['I3'] = parametros.get('tipo_medidor', '')

    # Linha 4: Data de Instalação
    sheet.merge_cells('C4:J4')
    sheet['A4'] = 'Data de Instalação'
    for col in ['A', 'B', 'C', 'D','E','F','G','H','I','J']:
        cell = sheet[f'{col}4']
        cell.border = thin_border

    # Adicionar valor se disponível
    # sheet['I4'] = parametros.get('data_instalacao', '')

    # Linha 5: Cabeçalhos de Seção
    sheet.merge_cells('A5:D5')
    sheet['A5'] = 'Leitura Equipamento'
    sheet['A5'].font = header_font
    sheet['A5'].alignment = center_alignment
    for col in ['A', 'B', 'C', 'D']:
        cell = sheet[f'{col}5']
        cell.border = thin_border
        cell.fill = fill_cinza

    sheet.merge_cells('E5:G5')
    sheet['E5'] = 'Resultados'
    sheet['E5'].font = header_font
    sheet['E5'].alignment = center_alignment
    for col in ['E', 'F', 'G']:
        cell = sheet[f'{col}5']
        cell.border = thin_border
        cell.fill = fill_cinza


    sheet.merge_cells('H5:J5')
    sheet['H5'] = 'Vazão Média - diária'
    sheet['H5'].font = header_font
    sheet['H5'].alignment = center_alignment
    for col in ['H', 'I', 'J']:
        cell = sheet[f'{col}5']
        cell.border = thin_border
        cell.fill = fill_cinza    
    

    # Linha 6: Cabeçalhos das Colunas de Dados
    data_headers = ['Data', 'Hora', 'Horimetro', 'Medidor de Vazão', 
                    'Tempo de Captação (h)', 'Volume diário (m3)', 
                    'Volume acumulado Mensal (m3)', 'Valor', 'Unidade', 'Observação']
    sheet.append(data_headers)
    for col_idx, header in enumerate(data_headers, 1):
        cell = sheet.cell(row=6, column=col_idx)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = fill_cinza

    # --- Escrever Dados --- 
    start_row = 7
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        sheet.append(list(row))
        # Aplicar formatação de número se necessário
        sheet.cell(row=r_idx, column=3).number_format = '0.000' # Horimetro
        sheet.cell(row=r_idx, column=4).number_format = '0.000' # Medidor de Vazão
        sheet.cell(row=r_idx, column=5).number_format = '0.00'  # Tempo de Captação
        sheet.cell(row=r_idx, column=6).number_format = '0.000' # Volume diário
        sheet.cell(row=r_idx, column=7).number_format = '0.000' # Volume acumulado
        sheet.cell(row=r_idx, column=8).number_format = '0.00'  # Valor (Vazão Média)

        for c_idx in range(1, len(row) + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.border = thin_border



    # --- Ajustar Largura das Colunas --- 
    for col_idx, header in enumerate(data_headers, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(header) + 2 # Padding
        for row_idx in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                 # Considerar formatação de número ao calcular largura
                 if isinstance(cell_value, (int, float)):
                      cell_str = f"{cell_value:.3f}" # Ajustar formato conforme a coluna
                 else:
                      cell_str = str(cell_value)                      
                 max_length = max(max_length, len(cell_str) + 2)
        sheet.column_dimensions[column_letter].width = max_length

    # Salvar o arquivo em memória
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

